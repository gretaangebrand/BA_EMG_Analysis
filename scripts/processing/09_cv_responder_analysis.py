"""
09_cv_responder_analysis.py
============================
Deskriptive Zusatzanalyse: Variationskoeffizient (CV) der EMG-Aktivierung
über die drei Zyklusphasen (PER/OVU/LUT) pro Probandin × Muskel × Übung.

Zweck:
  Ergänzt die inferenzstatistische Friedman-Analyse um ein individuelles
  Streuungsmaß. Der CV gibt an, wie stark eine einzelne Probandin
  zyklusabhängig in ihrer Muskelaktivierung schwankt.

Definition:
  CV [%] = (SD / Mittelwert) × 100
  berechnet über die drei Phasenwerte einer Person × Muskel × Übung.

Klassifikation:
  - Responder     : CV > 15 %  (deutliche zyklusabhängige Variation)
  - Non-Responder : CV ≤ 15 %  (stabile Aktivierung über den Zyklus)

Abschnitte pro Übungstyp:
  - SQ  → "Gesamt"  (keine Events definiert)
  - CMJ → "Landung" (IC → maximale Kniebeugung)
  - DJ  → "Landung2" (zweite Landung nach reaktivem Absprung)

Die Abschnitt-Zuordnung ist konsistent mit Script 08 (statistische
Auswertung). Der "Gesamt"-Abschnitt wird bei CMJ und DJ bewusst
nicht verwendet, um den Fokus auf die biomechanisch relevante
Landungsphase zu legen.

Kennwerte:
  - mean_emg
  - peak_emg

Voraussetzung:
  Nur Probandinnen mit vollständigen Daten in allen 3 Phasen werden einbezogen
  (konsistente Berechnung).

Output:
  - cv_responder_individuell.csv       (alle CVs pro Person × Muskel × Übung)
  - cv_responder_gruppenmittel.csv     (Gruppen-Median des CV pro Muskel × Übung)
  - cv_responder_uebersicht.xlsx       (formatierte Excel-Ausgabe)
  - cv_responder_heatmap_*.svg         (eine Heatmap pro Kennwert)
  - cv_responder_latex.tex             (LaTeX-Tabelle für Thesis)
"""

from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap

import sys
sys.path.insert(0, str(Path(r'C:\Users\Greta\OneDrive\Desktop\MCI\3-SS2026\BA\BA_Daten_EMG')))

from scripts.utils.config import MUSCLE_NAMES

# ============================================================
# PFADE
# ============================================================
FEATURES_CSV = Path(r"C:\Users\Greta\OneDrive\Desktop\MCI\3-SS2026\BA\BA_Daten_EMG\data\06_emg_features\emg_features_statistic.csv")
OUTPUT_DIR   = Path(r"C:\Users\Greta\OneDrive\Desktop\MCI\3-SS2026\BA\BA_Daten_EMG\outputs\cv_responder")

# ============================================================
# EINSTELLUNGEN
# ============================================================
PHASE_ORDER = ["PER", "OVU", "LUT"]
KENNWERTE   = ["mean_emg", "peak_emg"]
CV_THRESHOLD = 15.0   # % — Grenze Responder/Non-Responder

# Abschnitts-Zuordnung pro Übungstyp
ABSCHNITT_PRO_UEBUNG = {
    "SQ bilateral":    "Gesamt",
    "SQ einbeinig R":  "Gesamt",
    "CMJ bilateral":   "Landung",
    "CMJ einbeinig R": "Landung",
    "DJ bilateral":    "Landung2",
}


# ============================================================
# CV-BERECHNUNG
# ============================================================

def compute_cv(values: list[float]) -> float:
    """
    Variationskoeffizient in Prozent.
    CV = (SD / Mittelwert) × 100
    SD mit ddof=1 (Stichproben-SD).
    """
    arr = np.array(values, dtype=float)
    arr = arr[~np.isnan(arr)]
    if len(arr) < 2:
        return np.nan
    mean = np.mean(arr)
    if mean == 0 or np.isclose(mean, 0):
        return np.nan
    sd = np.std(arr, ddof=1)
    return (sd / abs(mean)) * 100.0


def classify_responder(cv: float) -> str:
    """Klassifikation als Responder / Non-Responder / n.a."""
    if pd.isna(cv):
        return "n.a."
    return "Responder" if cv > CV_THRESHOLD else "Non-Responder"


# ============================================================
# HAUPT-ANALYSE
# ============================================================

def build_cv_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Berechnet CV pro Probandin × Übung × Muskel × Kennwert.
    Nur wenn alle 3 Phasen vorhanden sind.
    """
    records = []

    for uebung, abschnitt in ABSCHNITT_PRO_UEBUNG.items():
        sub_u = df[(df["Uebung"] == uebung) & (df["Abschnitt"] == abschnitt)]
        if sub_u.empty:
            print(f"  [INFO] Keine Daten für {uebung} / {abschnitt}")
            continue

        for muskel in MUSCLE_NAMES:
            sub_m = sub_u[sub_u["Muskel"] == muskel]
            if sub_m.empty:
                continue

            for subj in sorted(sub_m["Subject"].unique()):
                sub_s = sub_m[sub_m["Subject"] == subj]

                for kennwert in KENNWERTE:
                    # Werte pro Phase holen
                    phase_vals = {}
                    for phase in PHASE_ORDER:
                        phase_row = sub_s[sub_s["Phase"] == phase]
                        if not phase_row.empty:
                            val = phase_row[kennwert].values[0]
                            if pd.notna(val):
                                phase_vals[phase] = float(val)

                    # Nur wenn alle 3 Phasen vorhanden
                    if len(phase_vals) < 3:
                        continue

                    vals = [phase_vals[p] for p in PHASE_ORDER]
                    mean_val = np.mean(vals)
                    sd_val   = np.std(vals, ddof=1)
                    cv_val   = compute_cv(vals)

                    records.append({
                        "Subject":    subj,
                        "Uebung":     uebung,
                        "Abschnitt":  abschnitt,
                        "Muskel":     muskel,
                        "Kennwert":   kennwert,
                        "PER":        phase_vals["PER"],
                        "OVU":        phase_vals["OVU"],
                        "LUT":        phase_vals["LUT"],
                        "Mittelwert": mean_val,
                        "SD":         sd_val,
                        "CV_%":       cv_val,
                        "Responder":  classify_responder(cv_val),
                    })

    return pd.DataFrame(records)


def build_group_summary(df_cv: pd.DataFrame) -> pd.DataFrame:
    """
    Gruppen-Zusammenfassung pro Übung × Muskel × Kennwert:
      - Median, Min, Max des individuellen CV
      - Anzahl Responder / Non-Responder
    """
    records = []
    for (uebung, abschnitt, muskel, kennwert), grp in df_cv.groupby(
        ["Uebung", "Abschnitt", "Muskel", "Kennwert"]
    ):
        cv_vals = grp["CV_%"].dropna()
        if cv_vals.empty:
            continue

        n_total   = len(grp)
        n_resp    = (grp["Responder"] == "Responder").sum()
        n_nonresp = (grp["Responder"] == "Non-Responder").sum()

        records.append({
            "Uebung":          uebung,
            "Abschnitt":       abschnitt,
            "Muskel":          muskel,
            "Kennwert":        kennwert,
            "n_gesamt":        n_total,
            "n_Responder":     n_resp,
            "n_NonResponder":  n_nonresp,
            "Responder_%":     (n_resp / n_total * 100) if n_total > 0 else np.nan,
            "CV_median":       cv_vals.median(),
            "CV_mean":         cv_vals.mean(),
            "CV_min":          cv_vals.min(),
            "CV_max":          cv_vals.max(),
        })

    return pd.DataFrame(records)


# ============================================================
# VISUALISIERUNG: HEATMAP
# ============================================================

def _create_heatmap(df_cv: pd.DataFrame, kennwert: str, out_path: Path):
    """
    Heatmap: Zeilen = Probandinnen, Spalten = Muskel × Übung.
    Zellwert = CV %, Farbe = Responder-Status.
    """
    sub = df_cv[df_cv["Kennwert"] == kennwert].copy()
    if sub.empty:
        print(f"  [INFO] Keine CV-Daten für {kennwert}, kein Plot erzeugt.")
        return

    # Pivot: Subject × (Uebung_Muskel)
    sub["Spalte"] = sub["Uebung"] + " | " + sub["Muskel"]

    # Spaltenreihenfolge: pro Übung dann Muskel
    spalten_order = []
    for uebung in ABSCHNITT_PRO_UEBUNG.keys():
        for muskel in MUSCLE_NAMES:
            col = f"{uebung} | {muskel}"
            if col in sub["Spalte"].values:
                spalten_order.append(col)

    pivot = sub.pivot_table(
        index="Subject", columns="Spalte", values="CV_%", aggfunc="first"
    )
    pivot = pivot.reindex(columns=spalten_order)

    subjects = sorted(pivot.index.tolist())
    pivot = pivot.reindex(index=subjects)

    n_rows = len(pivot)
    n_cols = len(pivot.columns)

    # Größere Figur für bessere Lesbarkeit
    fig_w = max(16, 0.7 * n_cols + 4)
    fig_h = max(8, 0.55 * n_rows + 3)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))

    # Diskrete Colormap: rot für Responder (>15), grau/grün für Non-Responder
    cmap = LinearSegmentedColormap.from_list(
        "cv_cmap",
        ["#2E7D32", "#A5D6A7", "#FFF59D", "#FFB74D", "#E53935"],
        N=256,
    )

    data = pivot.values
    vmax = max(30, np.nanmax(data) if not np.all(np.isnan(data)) else 30)

    im = ax.imshow(data, cmap=cmap, aspect="auto", vmin=0, vmax=vmax)

    # Keine Zahlen in Zelle - nur Farbe
    """
    for i in range(n_rows):
        for j in range(n_cols):
            val = data[i, j]
            if pd.notna(val):
                color = "white" if val > vmax * 0.6 else "black"
                fw = "bold" if val > CV_THRESHOLD else "normal"
                ax.text(j, i, f"{val:.1f}", ha="center", va="center",
                        fontsize=7, color=color, fontweight=fw)
            else:
                ax.text(j, i, "–", ha="center", va="center",
                        fontsize=7, color="#bbbbbb")
                        """

    # Achsen mit größerer Schrift
    ax.set_xticks(range(n_cols))
    ax.set_xticklabels(pivot.columns, rotation=45, ha="right", fontsize=12)
    ax.set_yticks(range(n_rows))
    ax.set_yticklabels(subjects, fontsize=12)
    ax.set_xlabel("Übung | Muskel", fontsize=13, labelpad=8)
    ax.set_ylabel("Probandin", fontsize=13)

    # Gitternetz zwischen Übungen (nach jedem 5. Muskel)
    for x in range(5, n_cols, 5):
        ax.axvline(x - 0.5, color="white", linewidth=1.5)

    # Grenzlinie bei CV = 15 (in Colorbar markieren)
    cbar = plt.colorbar(im, ax=ax, shrink=0.7, pad=0.02)
    cbar.set_label("CV in %", fontsize=11)
    cbar.ax.tick_params(labelsize=9)
    cbar.ax.axhline(CV_THRESHOLD, color="black", linewidth=1.2,
                    linestyle="--")
    cbar.ax.text(1.5, CV_THRESHOLD, f"  Schwelle {CV_THRESHOLD:.0f}%",
                 va="center", fontsize=9)

    #ax.set_title(
        #f"CV-Responder-Analyse – {kennwert}   "
        #f"(Responder: CV > {CV_THRESHOLD:.0f} %)",
        #fontsize=14, fontweight="bold", pad=12,
    #)

    plt.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  SVG (Heatmap {kennwert}): {out_path.name}")


# ============================================================
# EXCEL-AUSGABE
# ============================================================

def _save_excel(df_cv: pd.DataFrame, df_group: pd.DataFrame, out_path: Path):
    """Formatierte Excel-Übersicht mit zwei Sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Individuelle CVs ──
    ws1 = wb.active
    ws1.title = "Individuell"

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    responder_fill = PatternFill("solid", fgColor="FFE0E0")
    nonresp_fill = PatternFill("solid", fgColor="E0F0E0")
    normal_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    cols_individ = ["Subject", "Uebung", "Abschnitt", "Muskel", "Kennwert",
                    "PER", "OVU", "LUT", "Mittelwert", "SD", "CV_%", "Responder"]
    widths_individ = [10, 20, 14, 24, 12, 10, 10, 10, 12, 10, 10, 16]

    # Header
    for c, h in enumerate(cols_individ, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(c)].width = widths_individ[c - 1]

    # Daten
    for r_idx, (_, row) in enumerate(df_cv.iterrows(), 2):
        for c_idx, col in enumerate(cols_individ, 1):
            val = row[col]
            if isinstance(val, float) and not pd.isna(val):
                val = round(val, 2)
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font = normal_font
            cell.alignment = center
            cell.border = thin_border

            if col == "Responder":
                if val == "Responder":
                    cell.fill = responder_fill
                    cell.font = Font(name="Arial", size=10, bold=True,
                                     color="C62828")
                elif val == "Non-Responder":
                    cell.fill = nonresp_fill
                    cell.font = Font(name="Arial", size=10,
                                     color="2E7D32")

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Gruppenmittel ──
    ws2 = wb.create_sheet("Gruppenmittel")

    cols_group = ["Uebung", "Abschnitt", "Muskel", "Kennwert",
                  "n_gesamt", "n_Responder", "n_NonResponder", "Responder_%",
                  "CV_median", "CV_mean", "CV_min", "CV_max"]
    widths_group = [20, 14, 24, 12, 10, 12, 14, 14, 12, 12, 10, 10]

    for c, h in enumerate(cols_group, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(c)].width = widths_group[c - 1]

    for r_idx, (_, row) in enumerate(df_group.iterrows(), 2):
        for c_idx, col in enumerate(cols_group, 1):
            val = row[col]
            if isinstance(val, float) and not pd.isna(val):
                val = round(val, 2)
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = normal_font
            cell.alignment = center
            cell.border = thin_border

    ws2.freeze_panes = "A2"

    wb.save(out_path)
    print(f"  Excel: {out_path.name}")


# ============================================================
# LATEX-TABELLE
# ============================================================

def _generate_latex_table(df_group: pd.DataFrame, out_path: Path):
    """
    Kompakte LaTeX-Tabelle: pro Übung × Muskel × Kennwert
    Median-CV und Anzahl Responder.
    """
    lines = [
        "% Automatisch generiert von 09_cv_responder_analysis.py",
        "% CV-Responder-Analyse - Streuung der EMG-Aktivierung über Zyklusphasen",
        "",
    ]

    for kennwert in KENNWERTE:
        sub = df_group[df_group["Kennwert"] == kennwert]
        if sub.empty:
            continue

        kw_label = "mean" if kennwert == "mean_emg" else "peak"
        caption = (f"CV-Responder-Analyse – {kw_label}_emg. "
                   f"Median-CV und Anzahl Responder (CV > {CV_THRESHOLD:.0f}\\,\\%) "
                   f"pro Übung und Muskel.")

        lines.append(r"\begin{table}[H]")
        lines.append(r"  \centering")
        lines.append(f"  \\caption{{{caption}}}")
        lines.append(f"  \\label{{tab:cv_responder_{kw_label}}}")
        lines.append(r"  \small")
        lines.append(r"  \begin{tabular}{l l c c c c}")
        lines.append(r"    \hline")
        lines.append(
            r"    \textbf{Übung} & \textbf{Muskel} & "
            r"\textbf{CV-Median / \%]} & \textbf{CV-Range / \%]} & "
            r"\textbf{Responder} & \textbf{n} \\"
        )
        lines.append(r"    \hline")

        # Sortierung: Übung dann Muskel
        uebung_order = list(ABSCHNITT_PRO_UEBUNG.keys())
        for uebung in uebung_order:
            sub_u = sub[sub["Uebung"] == uebung]
            if sub_u.empty:
                continue

            first_row = True
            for muskel in MUSCLE_NAMES:
                row = sub_u[sub_u["Muskel"] == muskel]
                if row.empty:
                    continue
                r = row.iloc[0]

                uebung_str = uebung if first_row else ""
                first_row = False

                median = f"{r['CV_median']:.1f}".replace(".", "{,}")
                cv_range = (f"{r['CV_min']:.1f}--{r['CV_max']:.1f}"
                            .replace(".", "{,}"))
                resp_str = f"{r['n_Responder']}/{r['n_gesamt']}"

                lines.append(
                    f"    {uebung_str} & M.~{muskel.lower()} & "
                    f"{median} & {cv_range} & "
                    f"{resp_str} & {r['n_gesamt']} \\\\"
                )
            lines.append(r"    \hline")

        lines.append(r"  \end{tabular}")
        lines.append(r"\end{table}")
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  LaTeX: {out_path.name}")


# ============================================================
# HAUPTFUNKTION
# ============================================================

def main():
    print("=" * 70)
    print("09_cv_responder_analysis.py  –  CV-Responder-Analyse")
    print("=" * 70)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if not FEATURES_CSV.exists():
        print(f"[FEHLER] Features-Datei nicht gefunden:\n  {FEATURES_CSV}")
        return

    df = pd.read_csv(FEATURES_CSV)
    print(f"\nFeatures geladen: {len(df)} Zeilen")

    # Kennwerte prüfen
    for kw in KENNWERTE:
        if kw not in df.columns:
            print(f"[FEHLER] Spalte '{kw}' nicht in Features-CSV vorhanden.")
            print(f"  Verfügbare Spalten: {list(df.columns)}")
            return

    # ── Abschnitt-Filter prüfen ──
    # Die CV-Analyse nutzt pro Übung genau einen Abschnitt (siehe
    # ABSCHNITT_PRO_UEBUNG). Wir prüfen vorab, ob für diese Kombinationen
    # überhaupt Daten vorhanden sind, damit bei falscher Konfiguration
    # ein klarer Fehler statt eines leisen Leerlaufs entsteht.
    print(f"\nAbschnitt-Konfiguration:")
    for uebung, abschnitt in ABSCHNITT_PRO_UEBUNG.items():
        print(f"  {uebung:22s} -> {abschnitt}")

    masks = []
    for uebung, abschnitt in ABSCHNITT_PRO_UEBUNG.items():
        masks.append(
            (df["Uebung"] == uebung) & (df["Abschnitt"] == abschnitt)
        )
    combined = masks[0]
    for m in masks[1:]:
        combined = combined | m
    n_matched = combined.sum()

    print(f"  Zeilen in CSV gesamt: {len(df)}")
    print(f"  Zeilen passend zum Filter: {n_matched}")

    if n_matched == 0:
        print(f"\n[FEHLER] Keine Zeilen passen zur Filter-Konfiguration.")
        print(f"  Moegliche Ursachen:")
        print(f"    - Uebungs-Namen in ABSCHNITT_PRO_UEBUNG stimmen nicht")
        print(f"      mit den Werten in der Features-CSV ueberein")
        print(f"    - Abschnitt-Namen sind falsch geschrieben")
        print(f"  Gefundene Uebungen in CSV: "
              f"{sorted(df['Uebung'].unique())}")
        print(f"  Gefundene Abschnitte in CSV: "
              f"{sorted(df['Abschnitt'].unique())}")
        return

    # Warnung: Uebungen in CSV, die nicht in der Konfiguration stehen
    uebungen_csv = set(df["Uebung"].unique())
    uebungen_cfg = set(ABSCHNITT_PRO_UEBUNG.keys())
    fehlend = uebungen_csv - uebungen_cfg
    if fehlend:
        print(f"\n[WARNUNG] Folgende Uebungen sind in der CSV vorhanden,")
        print(f"          aber nicht in ABSCHNITT_PRO_UEBUNG und werden")
        print(f"          deshalb von der CV-Analyse ausgeschlossen:")
        for u in sorted(fehlend):
            print(f"  - {u}")

    # ── CV pro Person × Muskel × Übung × Kennwert ──
    print("\nBerechne CVs...")
    df_cv = build_cv_table(df)

    if df_cv.empty:
        print("[WARNUNG] Keine CV-Datensätze erzeugt.")
        return

    print(f"  Erzeugte CV-Datensätze: {len(df_cv)}")

    # ── Gruppen-Zusammenfassung ──
    df_group = build_group_summary(df_cv)

    # ── CSV-Ausgabe ──
    csv_individ = OUTPUT_DIR / "cv_responder_individuell.csv"
    csv_group   = OUTPUT_DIR / "cv_responder_gruppenmittel.csv"
    df_cv.to_csv(csv_individ, index=False)
    df_group.to_csv(csv_group, index=False)
    print(f"\n  CSV (Individuell)  : {csv_individ.name}")
    print(f"  CSV (Gruppenmittel): {csv_group.name}")

    # ── Excel ──
    _save_excel(df_cv, df_group, OUTPUT_DIR / "cv_responder_uebersicht.xlsx")

    # ── Heatmaps (eine pro Kennwert) ──
    for kw in KENNWERTE:
        _create_heatmap(
            df_cv, kw,
            OUTPUT_DIR / f"cv_responder_heatmap_{kw}.svg"
        )

    # ── LaTeX-Tabelle ──
    _generate_latex_table(df_group, OUTPUT_DIR / "cv_responder_latex.tex")

    # ── Zusammenfassung Konsole ──
    print(f"\n{'=' * 70}")
    print("ZUSAMMENFASSUNG")
    print(f"{'=' * 70}")
    for kw in KENNWERTE:
        sub = df_cv[df_cv["Kennwert"] == kw]
        if sub.empty:
            continue
        n_total = len(sub)
        n_resp  = (sub["Responder"] == "Responder").sum()
        pct     = n_resp / n_total * 100 if n_total > 0 else 0
        cv_med  = sub["CV_%"].median()
        print(f"\n  Kennwert: {kw}")
        print(f"    CV-Median (alle):       {cv_med:.1f} %")
        print(f"    Responder (CV > {CV_THRESHOLD:.0f}%):    "
              f"{n_resp}/{n_total} ({pct:.1f} %)")

        # Pro Übung kurz auflisten
        for uebung in ABSCHNITT_PRO_UEBUNG.keys():
            sub_u = sub[sub["Uebung"] == uebung]
            if sub_u.empty:
                continue
            n_u = len(sub_u)
            n_r = (sub_u["Responder"] == "Responder").sum()
            print(f"      {uebung:20s} → {n_r}/{n_u} Responder, "
                  f"Median CV = {sub_u['CV_%'].median():.1f} %")

    print(f"\n{'=' * 70}")
    print(f"FERTIG – Output: {OUTPUT_DIR}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()