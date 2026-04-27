"""
08_statistical_analysis.py
===========================
Führt die komplette statistische Auswertung der EMG-Features durch.
 
Für jede Kombination aus Übung × Muskel × Abschnitt × Kennwert:
  1. Deskriptive Statistik (M ± SD pro Phase)
  2. Friedman-Test (non-parametrisch, für alle Kombinationen einheitlich)
  3. Kendalls W als Effektgröße
  4. Post-hoc: Paarweiser Wilcoxon-Vorzeichen-Rangtest mit Bonferroni-Korrektur

Begründung für einheitliche Friedman-Anwendung:
  - EMG-Kennwerte sind bei kleinen Stichproben (n ≤ 11) selten sicher
    normalverteilt.
  - Einheitliches Vorgehen erleichtert die Ergebnisinterpretation.
  - Friedman ist konservativer und damit robuster gegenüber Verletzungen
    der Normalitätsannahme.

Output:
  - statistische_ergebnisse.csv
  - statistische_ergebnisse.xlsx (formatiert)
  - latex_tabellen.tex (fertige LaTeX-Tabellen)
"""

from pathlib import Path
import pandas as pd
import numpy as np
from scipy import stats
import pingouin as pg
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)

# ============================================================
# PFADE
# ============================================================
FEATURES_CSV = Path(r"C:\Users\Greta\OneDrive\Desktop\MCI\3-SS2026\BA\BA_Daten_EMG\data\06_emg_features\emg_features_statistic.csv")
OUTPUT_DIR   = Path(r"C:\Users\Greta\OneDrive\Desktop\MCI\3-SS2026\BA\BA_Daten_EMG\outputs\statistics")
#REPORT_PATH  = Path(r"C:\Users\Greta\OneDrive\Desktop\MCI\3-SS2026\BA\BA_Daten_EMG\data\Pipeline_Reports.xlsx")

# ============================================================
# EINSTELLUNGEN
# ============================================================
ALPHA = 0.05
PHASE_ORDER = ["PER", "OVU", "LUT"]
KENNWERTE = ["mean_emg", "peak_emg"]

# ── Abschnitt-Filter ──────────────────────────────────────────
# Legt fest, welche Abschnitte pro Uebung in die finale Auswertung
# eingehen. Methodischer Hintergrund:
#
#   - CMJ: Nur die Landungsphase wird analysiert (vom initialen
#     Bodenkontakt bis zur maximalen Kniebeugung).
#   - DJ:  Nur die zweite Landungsphase wird analysiert (reaktive
#     Landung nach dem Absprung von der Box).
#   - SQ:  Der gesamte Bewegungszyklus wird analysiert, da keine
#     Event-Markierungen gesetzt sind.
#
# Der "Gesamt"-Abschnitt fuer CMJ und DJ bleibt in den Feature-
# Daten erhalten (fuer Sanity-Checks), geht aber nicht in die
# finale Statistik ein.
#
# Zum Deaktivieren des Filters (z.B. fuer Validierung): auf None
# setzen.
# --------------------------------------------------------------
ABSCHNITT_FILTER = {
    "CMJ bilateral":   ["Landung"],
    "CMJ einbeinig R": ["Landung"],
    "DJ bilateral":    ["Landung2"],
    "SQ bilateral":    ["Gesamt"],
    "SQ einbeinig R":  ["Gesamt"],
}
 
 
def interpret_kendalls_w(w):
    """Interpretation von Kendalls W nach Landis & Koch (1977)."""
    if pd.isna(w):
        return ""
    if w >= 0.7:
        return "stark"
    if w >= 0.4:
        return "moderat"
    if w >= 0.1:
        return "schwach"
    return "vernachlaessigbar"
 
 
def fmt_msd(m, sd):
    """Formatiert M ± SD mit 3 Nachkommastellen, Dezimalkomma."""
    if pd.isna(m) or pd.isna(sd):
        return "--"
    return f"{m:.2f} ± {sd:.2f}".replace(".", ",")
 
 
def fmt_p(p):
    """Formatiert p-Wert mit 3 Nachkommastellen."""
    if pd.isna(p):
        return ""
    if p < 0.001:
        return "<0,001"
    return f"{p:.2f}".replace(".", ",")
 
 
def fmt_num(val, decimals=3):
    """Formatiert eine Zahl mit Dezimalkomma."""
    if pd.isna(val):
        return ""
    return f"{val:.{decimals}f}".replace(".", ",")
 
 
# ============================================================
# STATISTISCHE ANALYSE PRO KOMBINATION
# ============================================================
 
def analyse_combination(df_long: pd.DataFrame, kennwert: str) -> dict:
    """
    Führt Friedman-Test + Post-hoc (Wilcoxon, Bonferroni) durch.
    """
    result = {}
 
    # ── Deskriptive Statistik ──
    for phase in PHASE_ORDER:
        vals = df_long[df_long["Phase"] == phase][kennwert].values
        if len(vals) > 0:
            result[f"M_{phase}"] = np.mean(vals)
            result[f"SD_{phase}"] = np.std(vals, ddof=1)
            result[f"n_{phase}"] = len(vals)
        else:
            result[f"M_{phase}"] = np.nan
            result[f"SD_{phase}"] = np.nan
            result[f"n_{phase}"] = 0
 
    # ── Vollständige Fälle prüfen ──
    subjects_per_phase = {
        p: set(df_long[df_long["Phase"] == p]["Subject"].values)
        for p in PHASE_ORDER
    }
    complete_subjects = (
        subjects_per_phase["PER"]
        & subjects_per_phase["OVU"]
        & subjects_per_phase["LUT"]
    )
 
    if len(complete_subjects) < 3:
        result["Test"] = "nicht durchführbar"
        result["Anmerkungen"] = (
            f"Nur {len(complete_subjects)} Probandinnen "
            f"mit Daten in allen 3 Phasen"
        )
        return result
 
    df_complete = df_long[df_long["Subject"].isin(complete_subjects)].copy()
    result["n_komplett"] = len(complete_subjects)
 
    # ── Prüfen ob Varianz vorhanden ──
    for phase in PHASE_ORDER:
        vals = df_complete[df_complete["Phase"] == phase][kennwert].values
        if np.std(vals, ddof=1) == 0:
            result["Test"] = "nicht durchführbar"
            result["Anmerkungen"] = f"Keine Varianz in Phase {phase}"
            return result
 
    # ── Friedman-Test ──
    result["Test"] = "Friedman"
 
    vals_per_phase = []
    for phase in PHASE_ORDER:
        phase_vals = (
            df_complete[df_complete["Phase"] == phase]
            .sort_values("Subject")[kennwert]
            .values
        )
        vals_per_phase.append(phase_vals)
 
    try:
        stat_fr, p_fr = stats.friedmanchisquare(*vals_per_phase)
        result["Chi2"] = stat_fr
        result["Teststatistik"] = fmt_num(stat_fr)
        result["df"] = str(len(PHASE_ORDER) - 1)
        result["p_wert"] = p_fr
        result["Signifikant"] = "ja" if p_fr < ALPHA else "nein"
 
        # Kendalls W
        n = len(vals_per_phase[0])
        k = len(PHASE_ORDER)
        result["Kendalls_W"] = (
            stat_fr / (n * (k - 1)) if n * (k - 1) > 0 else np.nan
        )
        result["W_interpretation"] = interpret_kendalls_w(result["Kendalls_W"])
 
    except Exception as e:
        result["Test"] = "Friedman fehlgeschlagen"
        result["Anmerkungen"] = str(e)
        return result
 
    # ── Post-hoc: Paarweiser Wilcoxon-Vorzeichen-Rangtest mit Bonferroni ──
    posthoc_p_values = []
    if result.get("Signifikant") == "ja":
        pairs = [("PER", "OVU"), ("PER", "LUT"), ("OVU", "LUT")]
        n_pairs = len(pairs)
 
        for p1, p2 in pairs:
            v1 = (
                df_complete[df_complete["Phase"] == p1]
                .sort_values("Subject")[kennwert].values
            )
            v2 = (
                df_complete[df_complete["Phase"] == p2]
                .sort_values("Subject")[kennwert].values
            )
            try:
                _, p_w = stats.wilcoxon(v1, v2)
                # Bonferroni-Korrektur: p * Anzahl Paare, gedeckelt bei 1
                p_corrected = min(p_w * n_pairs, 1.0)
                result[f"posthoc_{p1}_{p2}_p"] = p_corrected
                posthoc_p_values.append(p_corrected)
            except Exception:
                result[f"posthoc_{p1}_{p2}_p"] = np.nan
        
        # Kleinster Post-hoc p-Wert
        if posthoc_p_values:
            result["p_posthoc_min"] = min(posthoc_p_values)
        else:
            result["p_posthoc_min"] = np.nan
    else:
        result["p_posthoc_min"] = np.nan
 
    return result
 
 
# ============================================================
# HAUPTFUNKTION
# ============================================================
 
def main():
    print("=" * 70)
    print("08_statistical_analysis.py  –  Friedman-basierte Auswertung")
    print("=" * 70)
 
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
 
    if not FEATURES_CSV.exists():
        print(f"[FEHLER] Datei nicht gefunden:\n  {FEATURES_CSV}")
        return
 
    df = pd.read_csv(FEATURES_CSV)
    print(f"\nFeatures geladen: {len(df)} Zeilen")
    print(f"  Subjects:   {sorted(df['Subject'].unique())}")

    # ── Abschnitt-Filter anwenden ──
    if ABSCHNITT_FILTER is not None:
        n_vor = len(df)

        masks = []
        for uebung, allowed_abschnitte in ABSCHNITT_FILTER.items():
            masks.append(
                (df["Uebung"] == uebung)
                & (df["Abschnitt"].isin(allowed_abschnitte))
            )

        # OR-Verknuepfung aller Einzelmasken
        combined_mask = masks[0]
        for m in masks[1:]:
            combined_mask = combined_mask | m

        df = df[combined_mask].reset_index(drop=True)
        n_nach = len(df)

        print(f"\nAbschnitt-Filter aktiv:")
        for uebung, abs_list in ABSCHNITT_FILTER.items():
            print(f"  {uebung:22s} -> {abs_list}")
        print(f"  Zeilen vor Filter:  {n_vor}")
        print(f"  Zeilen nach Filter: {n_nach}")

        # Safety-Check: ist nach dem Filter noch etwas da?
        if n_nach == 0:
            print(f"\n[FEHLER] Nach Anwendung des Abschnitt-Filters keine "
                  f"Daten mehr vorhanden.")
            print(f"  Moegliche Ursachen:")
            print(f"    - ABSCHNITT_FILTER enthaelt Uebungs- oder Abschnitts-"
                  f"namen, die nicht in den Daten vorkommen")
            print(f"    - Verfuegbare Uebungen in den Daten:")
            for u in sorted(pd.read_csv(FEATURES_CSV)['Uebung'].unique()):
                print(f"        '{u}'")
            print(f"    - Verfuegbare Abschnitte in den Daten:")
            for a in sorted(pd.read_csv(FEATURES_CSV)['Abschnitt'].unique()):
                print(f"        '{a}'")
            return

        # Zusatz-Warnung bei einzelnen leeren Kombinationen
        for uebung, allowed_abschnitte in ABSCHNITT_FILTER.items():
            n_uebung = len(df[df["Uebung"] == uebung])
            if n_uebung == 0:
                print(f"  [WARNUNG] Keine Daten fuer Uebung '{uebung}' "
                      f"mit Abschnitt(en) {allowed_abschnitte}")

    # ── Alle Kombinationen durchrechnen ──
    all_results = []
    n_friedman, n_sig, n_skipped = 0, 0, 0
 
    combinations = (
        df.groupby(["Uebung", "Muskel", "Abschnitt"])
        .size().reset_index()
    )
    total = len(combinations) * len(KENNWERTE)
    print(f"\n{total} Analysen werden durchgeführt...\n")
 
    for _, combo in combinations.iterrows():
        uebung    = combo["Uebung"]
        muskel    = combo["Muskel"]
        abschnitt = combo["Abschnitt"]
 
        mask = (
            (df["Uebung"] == uebung)
            & (df["Muskel"] == muskel)
            & (df["Abschnitt"] == abschnitt)
        )
        df_sub = df[mask].copy()
 
        for kennwert in KENNWERTE:
            result = analyse_combination(df_sub, kennwert)
            result["Uebung"] = uebung
            result["Muskel"] = muskel
            result["Abschnitt"] = abschnitt
            result["Kennwert"] = kennwert
            all_results.append(result)
 
            test = result.get("Test", "")
            if test == "Friedman":
                n_friedman += 1
            elif test in ("nicht durchführbar", "Friedman fehlgeschlagen"):
                n_skipped += 1
            if result.get("Signifikant") == "ja":
                n_sig += 1
 
            sig = " ***" if result.get("Signifikant") == "ja" else ""
            p_val = result.get("p_wert", np.nan)
            p_str = f"p={p_val:.4f}" if not pd.isna(p_val) else "p=n/a"
            print(
                f"  {uebung:22s} | {muskel:25s} | {abschnitt:12s} | "
                f"{kennwert:8s} | {test:12s} | {p_str}{sig}"
            )
 
    df_results = pd.DataFrame(all_results)
 
    # ── CSV speichern ──
    csv_out = OUTPUT_DIR / "statistische_ergebnisse.csv"
    df_results.to_csv(csv_out, index=False)
    print(f"\n  CSV: {csv_out}")
 
    # ── Formatierte Excel ──
    _save_formatted_excel(df_results, OUTPUT_DIR / "statistische_ergebnisse.xlsx")
 
    # ── LaTeX-Tabellen ──
    _generate_latex_tables(df_results, OUTPUT_DIR / "latex_tabellen.tex")
 
    # ── Zusammenfassung ──
    print(f"\n{'='*70}")
    print("ZUSAMMENFASSUNG")
    print(f"{'='*70}")
    print(f"  Analysen gesamt:       {len(df_results)}")
    print(f"  Friedman durchgeführt: {n_friedman}")
    print(f"  Nicht durchführbar:    {n_skipped}")
    print(f"  Signifikant:           {n_sig}")
 
    if n_sig > 0:
        print(f"\n  Signifikante Ergebnisse:")
        for _, r in df_results[df_results["Signifikant"] == "ja"].iterrows():
            print(f"    {r['Uebung']} | {r['Muskel']} | "
                  f"{r['Abschnitt']} | {r['Kennwert']} | p={r['p_wert']:.4f}")
    else:
        print(f"\n  Kein signifikanter Effekt gefunden.")
 
    # Kendalls W Übersicht
    if "Kendalls_W" in df_results.columns:
        w_vals = df_results["Kendalls_W"].dropna()
        if not w_vals.empty:
            print(f"\n  Effektgrössen (Kendalls W):")
            print(f"    Median:          {w_vals.median():.2f}")
            print(f"    Max:             {w_vals.max():.2f}")
            print(f"    Stark (≥0.7):    {(w_vals >= 0.7).sum()}")
            print(f"    Moderat (≥0.4):  {((w_vals >= 0.4) & (w_vals < 0.7)).sum()}")
            print(f"    Schwach (≥0.1):  {((w_vals >= 0.1) & (w_vals < 0.4)).sum()}")
 
    print(f"\n{'='*70}")
    print(f"FERTIG – Output: {OUTPUT_DIR}")
    print(f"{'='*70}")
 
 
# ============================================================
# EXCEL-AUSGABE
# ============================================================
 
def _save_formatted_excel(df: pd.DataFrame, out_path: Path):
    """
    Speichert die Ergebnisse im formatierten Excel-Layout.
    Spalten: deskriptive Kennwerte, Friedman-Ergebnis, Kendalls W, Post-hoc.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistik_Ergebnisse"
 
    # ── Styles ──
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    section_fill = PatternFill('solid', fgColor='D6E4F0')
    section_font = Font(bold=True, name='Arial', size=10, color='2F5496')
    normal_font = Font(name='Arial', size=10)
    sig_fill = PatternFill('solid', fgColor='E2EFDA')
    center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'),
    )
    zebra = PatternFill('solid', fgColor='F2F7FB')
 
    # ── Spalten-Header (ohne ANOVA-spezifische Spalten) ──
    headers = [
        "Übung", "Muskel", "Abschnitt", "Kennwert",
        "M ± SD\nPER (% BL)", "M ± SD\nOVU (% BL)", "M ± SD\nLUT (% BL)",
        "Test",
        "p-Friedman",
        "p-post-hoc\n(min)",
        "Kendalls W",
        "Signifikant?\n(p < 0,05)",
        "Post-hoc\nPER↔OVU (p)",
        "Post-hoc\nPER↔LUT (p)",
        "Post-hoc\nOVU↔LUT (p)",
        "Anmerkungen",
    ]
    col_widths = [20, 22, 16, 12, 18, 18, 18, 12, 12, 12, 12, 14, 14, 14, 14, 30]
 
    # ── Titel ──
    last_col_letter = chr(ord('A') + len(headers) - 1)
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'].value = "Statistische Ergebnisse – EMG-Kennwerte über die drei Zyklusphasen"
    ws['A1'].font = Font(bold=True, name='Arial', size=14, color='2F5496')
 
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'].value = ("Friedman-Test (non-parametrisch) | "
                      "Post-hoc: paarweiser Wilcoxon-Vorzeichen-Rangtest, "
                      "Bonferroni-korrigiert | α = 0,05")
    ws['A2'].font = Font(name='Arial', size=9, color='666666')
 
    ws.merge_cells(f'A3:{last_col_letter}3')
    ws['A3'].value = "Werte auf 3 Nachkommastellen gerundet"
    ws['A3'].font = Font(name='Arial', size=9, color='666666')
 
    # ── Spaltenbreiten ──
    for c, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c)].width = w
 
    # ── Abschnitte und Daten schreiben ──
    abschnitt_kennwert_order = [
        ("Gesamt", "mean_emg"),
        ("Gesamt", "peak_emg"),
        ("Landung", "mean_emg"),
        ("Landung", "peak_emg"),
        #("Bodenkontakt", "mean_emg"),
        #("Bodenkontakt", "peak_emg"),
        ("Landung2", "mean_emg"),
        ("Landung2", "peak_emg"),
    ]
 
    abschnitt_labels = {
        ("Landung",      "mean_emg"): "Landungsphase (CMJ) – Mean EMG",
        ("Landung",      "peak_emg"): "Landungsphase (CMJ) – Peak EMG",
        ("Landung2",     "mean_emg"): "Zweite Landungsphase (DJ) – Mean EMG",
        ("Landung2",     "peak_emg"): "Zweite Landungsphase (DJ) – Peak EMG",
        ("Gesamt",       "mean_emg"): "Gesamter Bewegungszyklus (Squat) – Mean EMG",
        ("Gesamt",       "peak_emg"): "Gesamter Bewegungszyklus (Squat) – Peak EMG",
        ("Bodenkontakt", "mean_emg"): "Bodenkontakt (DJ) – Mean EMG",
        ("Bodenkontakt", "peak_emg"): "Bodenkontakt (DJ) – Peak EMG",
    }
 
    current_row = 4
 
    for abschnitt, kennwert in abschnitt_kennwert_order:
        sub = df[
            (df["Abschnitt"] == abschnitt)
            & (df["Kennwert"] == kennwert)
        ]
        if sub.empty:
            continue
 
        # Abschnitts-Überschrift
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(headers),
        )
        cell = ws.cell(
            row=current_row, column=1,
            value=abschnitt_labels.get(
                (abschnitt, kennwert), f"{abschnitt} – {kennwert}"
            ),
        )
        cell.font = section_font
        cell.fill = section_fill
        current_row += 1
 
        # Header
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        current_row += 1
 
        # Datenzeilen
        sub_sorted = sub.sort_values(["Uebung", "Muskel"])
        for row_idx, (_, r) in enumerate(sub_sorted.iterrows()):
            values = [
                r.get("Uebung", ""),
                r.get("Muskel", ""),
                r.get("Abschnitt", ""),
                r.get("Kennwert", ""),
                fmt_msd(r.get("M_PER"), r.get("SD_PER")),
                fmt_msd(r.get("M_OVU"), r.get("SD_OVU")),
                fmt_msd(r.get("M_LUT"), r.get("SD_LUT")),
                r.get("Test", ""),
                fmt_p(r.get("p_wert")),  # p-Friedman
                fmt_p(r.get("p_posthoc_min")),  # kleinster Post-hoc p-Wert
                fmt_num(r.get("Kendalls_W")) if not pd.isna(r.get("Kendalls_W", np.nan)) else "",
                r.get("W_interpretation", ""),
                r.get("Signifikant", ""),
                fmt_p(r.get("posthoc_PER_OVU_p")),
                fmt_p(r.get("posthoc_PER_LUT_p")),
                fmt_p(r.get("posthoc_OVU_LUT_p")),
                r.get("Anmerkungen", ""),
            ]
 
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=c, value=val)
                cell.font = normal_font
                cell.alignment = center if c > 4 else left
                cell.border = thin_border
 
                # Zebrastreifen
                if row_idx % 2 == 1:
                    cell.fill = zebra
 
            # Signifikante Zeilen grün
            if r.get("Signifikant") == "ja":
                for c in range(1, len(headers) + 1):
                    ws.cell(row=current_row, column=c).fill = sig_fill
 
            ws.row_dimensions[current_row].height = 45
            current_row += 1
 
        # Leerzeile nach jedem Abschnitt
        current_row += 1
 
    ws.freeze_panes = 'A5'
    wb.save(out_path)
    print(f"  Excel: {out_path.name}")
 
 
# ============================================================
# LATEX-TABELLEN
# ============================================================
 
def _generate_latex_tables(df: pd.DataFrame, out_path: Path):
    """
    Erzeugt LaTeX-Tabellen pro Übung × Abschnitt.
    Spalten: Muskel | Kennwert | M±SD pro Phase | χ² | p | W
    """
    lines = [
        "% Automatisch generiert von 08_statistical_analysis.py",
        "% Friedman-Test | Post-hoc: Wilcoxon, Bonferroni",
        "",
    ]
 
    for (uebung, abschnitt), grp in df.groupby(["Uebung", "Abschnitt"]):
        label_safe = (
            uebung.replace(" ", "_").replace(".", "")
            + "_" + abschnitt
        )
        caption = f"EMG-Kennwerte – {uebung} ({abschnitt})"
 
        lines.append(r"\begin{table}[H]")
        lines.append(r"  \centering")
        lines.append(f"  \\caption{{{caption}}}")
        lines.append(f"  \\label{{tab:{label_safe}}}")
        lines.append(r"  \small")
        lines.append(r"  \begin{tabular}{l l c c c c c c }")
        lines.append(r"    \hline")
        lines.append(
            r"    \textbf{Muskel} & \textbf{Kennwert} & "
            r"\textbf{PER (\% BL)} & \textbf{OVU (\% BL)} & \textbf{LUT (\% BL)} & "
            r"\textbf{$p_{Friedman}$} & \textbf{$p_{post-hoc}$} & "
        )
        lines.append(r"    \hline")
 
        for muskel in grp["Muskel"].unique():
            first_row = True
            for kennwert in KENNWERTE:
                row = grp[
                    (grp["Muskel"] == muskel)
                    & (grp["Kennwert"] == kennwert)
                ]
                if row.empty:
                    continue
                r = row.iloc[0]
 
                muskel_str = (
                    f"M.~{muskel.lower()}" if first_row else ""
                )
                first_row = False
 
                def lfmt(phase):
                    m = r.get(f"M_{phase}", np.nan)
                    sd = r.get(f"SD_{phase}", np.nan)
                    if pd.isna(m):
                        return "--"
                    m_str = f"{m:.1f}".replace(".", "{,}")
                    sd_str = f"{sd:.1f}".replace(".", "{,}")
                    return f"{m_str} $\\pm$ {sd_str}"
 
                p_fr = r.get("p_wert", np.nan)
                if pd.isna(p_fr):
                    p_fr_str = "--"
                elif p_fr < 0.001:
                    p_fr_str = "$<$\\,0{,}001"
                else:
                    p_fr_str = f"{p_fr:.3f}".replace(".", "{,}")
                
                p_ph = r.get("p_posthoc_min", np.nan)
                if pd.isna(p_ph):
                    p_ph_str = "--"
                elif p_ph < 0.001:
                    p_ph_str = "$<$\\,0{,}001"
                else:
                    p_ph_str = f"{p_ph:.3f}".replace(".", "{,}")
 
                w_val = r.get("Kendalls_W", np.nan)
                w_str = (
                    f"{w_val:.2f}".replace(".", "{,}")
                    if not pd.isna(w_val) else "--"
                )
 
                kw = "mean" if kennwert == "mean_emg" else "peak"
 
                lines.append(
                    f"    {muskel_str} & {kw} & "
                    f"{lfmt('PER')} & {lfmt('OVU')} & {lfmt('LUT')} & "
                    f"{p_fr_str} & {p_ph_str}\\\\" #& {w_str} 
                )
            lines.append(r"    \hline")
 
        lines.append(r"  \end{tabular}")
        lines.append(r"\end{table}")
        lines.append("")
 
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  LaTeX: {out_path.name}")
 
 
if __name__ == "__main__":
    main()